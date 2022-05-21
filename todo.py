from openpyxl import load_workbook
from datetime import datetime
import pandas as pd

# excel path
path = 'utils/todo.xlsx'


def insert_todo(todo,mail):
    workbook = load_workbook(filename=path)
    spreadsheet = workbook.active
    spreadsheet.insert_rows(idx=2)
    spreadsheet["A2"] = todo
    spreadsheet["B2"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    spreadsheet["C2"] = "Pending"
    spreadsheet["D2"] = mail
    workbook.save(filename=path)


# todo show only pending tasks

def show_todo(mail):
    data = pd.read_excel(path, sheet_name='Sheet')
    data = data.loc[data['status'] == 'Pending']
    data = data.loc[data['mail'] == mail]
    return data


def finish_todo(identifier):
    workbook = load_workbook(filename=path)
    spreadsheet = workbook.active
    new_id = int(identifier) + 2
    if spreadsheet["C"+str(new_id)].value is not None:
        spreadsheet["C"+str(new_id)] = "Completed"
    workbook.save(filename=path)

def show_history(mail):
    data = pd.read_excel(path, sheet_name='Sheet')
    data = data.loc[data['mail'] == mail]
    return data
